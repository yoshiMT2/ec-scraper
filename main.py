from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import pandas as pd
import datetime
from amazon import amazon
from q10 import q10
from gome import gome
from suning import suning

def configure_driver():
    """Return chrome driver instance"""
    options = Options()
    preferences = {"profile.default_content_setting_values.notifあications" : 2}
    options.add_experimental_option("prefs", preferences)
    options.add_argument('--disable-software-rasterizer')
    options.add_argument("window-size=1920,1080")
    # options.add_argument('--headless')
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--ignore-ssl-errors')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    return driver


def main():
    print(datetime.datetime.now())

    driver = configure_driver()
    wait = WebDriverWait(driver, 10)

    COLUMN_NAMES = ['category', 'store_name', 'company_name', 'shipped_from',
                    'transporter', 'address', 'email', 'service_hour', 'tel',
                    'fax', 'product_url', 'open_date', 'office_hour', 'staff',
                    'staff_email', 'staff_tell', 'response']

    today = datetime.datetime.now().date()
    filename = f"{today}_セラー情報.xlsx"

    """ここからamazon"""
    wb = load_workbook('seller_info.xlsx')
    amazon_data1 = amazon(driver, wait, 'ビューティー')
    amazon_data2 = amazon(driver, wait, 'ファッション')
    amazon_data3 = amazon(driver, wait, 'ホーム＆キッチン')
    amazon_data = amazon_data1 + amazon_data2 + amazon_data3
    df = pd.DataFrame(amazon_data)
    df.columns = COLUMN_NAMES
    new_df = df.drop_duplicates(subset=['company_name'])
    data_list = new_df.to_numpy().tolist()
    ws = wb['amazon']
    for row in range(0, len(data_list)):
        for col in range(0, len(data_list[row])):
            ws.cell(row=row+4, column=col+2, value=data_list[row][col])
    wb.save(filename)

    """ここからq10"""
    print('Starting Q10')
    data_q10 = q10(driver)
    # wb = load_workbook(filename)
    #アマゾンを飛ばしてq10からプログラムを回す場合は、上の行をコメントアウトして、下の行をコメントイン
    wb = load_workbook('seller_info.xlsx')
    ws = wb['q10']
    for row in range(0, len(data_q10)):
        for col in range(0, len(data_q10[row])):
            ws.cell(row=row+4, column=col+2, value=data_q10[row][col])
    wb.save(filename)

    """ここからgome"""
    print('Starting Gome')
    data_gome = gome(driver)
    wb = load_workbook(filename)
    ws = wb['gome']
    for row in range(0, len(data_gome)):
        for col in range(0, len(data_gome[row])):
            ws.cell(row=row+4, column=col+2, value=data_gome[row][col])
    wb.save(filename)

    """ここからsuning"""
    print('Starting Suning')
    data_suning = suning(driver, wait)
    wb = load_workbook(filename)
    ws = wb['suning']
    for row in range(0, len(data_suning)):
        for col in range(0, len(data_suning[row])):
            ws.cell(row=row+4, column=col+2, value=data_suning[row][col])
    wb.save(filename)


if __name__ == '__main__':
    main()
